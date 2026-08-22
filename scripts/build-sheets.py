#!/usr/bin/env python3
"""
Utkarsh 2026 — the paper the day actually runs on.

Builds four Excel workbooks:

  1. gita-shloka-judging.xlsx    judges' scoring sheet
  2. vedic-quiz-marks.xlsx       roll + marks for both rounds
  3. devotional-essay-judging.xlsx
  4. registration-counter.xlsx   every student, every event, payment status

Run it with no arguments and you get the same four sheets with no names in
them — blank, numbered and printable, which is what you want if the laptop
dies and the whole morning goes to pen and paper.

With a CSV exported from the admin portal, the registered students are filled
in and the blank rows are added underneath for spot registrations:

    python3 scripts/build-sheets.py --csv ~/Downloads/registrations.csv

Every sheet is landscape, fits one page wide, repeats its header on every
printed page, and prints its gridlines — a judging sheet without gridlines is
useless the moment someone writes in it.
"""

import argparse, csv, datetime as dt, pathlib, re, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── House style ───────────────────────────────────────────────────────────────
SAFFRON   = 'C1440E'
INK       = '1F2933'
BAND      = 'FDF3E3'   # header fill
SPOT_FILL = 'F4F7FB'   # the blank spot-registration rows
RULE      = 'B7BEC7'

THIN = Side(style='thin', color=RULE)
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_F  = Font(name='Calibri', size=16, bold=True, color=SAFFRON)
SUB_F    = Font(name='Calibri', size=10, color='5A6672')
HEAD_F   = Font(name='Calibri', size=10, bold=True, color=INK)
CELL_F   = Font(name='Calibri', size=10, color=INK)
WRAP_MID = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_MID = Alignment(horizontal='left',   vertical='center', wrap_text=True)


class Col:
    """One column: its heading, width, and whether judges write in it."""
    def __init__(self, title, width=14, fill_in=False, align='left', note=None):
        self.title, self.width, self.fill_in = title, width, fill_in
        self.align, self.note = align, note


def parse_date(raw):
    """
    The admin CSV writes dates the way en-IN does: '3 Feb 2014'.

    September comes out as 'Sept', which is four letters and so does not match
    %b — every September birthday silently lost its age until this normalised
    it. Trailing full stops get the same treatment.
    """
    raw = (raw or '').strip().rstrip('.')
    if not raw or raw == '—':
        return None
    raw = re.sub(r'\bSept\b', 'Sep', raw, flags=re.I)
    for fmt in ('%d %b %Y', '%d %B %Y', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def age_on(dob, on):
    if not dob:
        return '?'
    return on.year - dob.year - ((on.month, on.day) < (dob.month, dob.day))


def build_sheet(wb, *, key, title, subtitle, cols, rows, spot_rows, first, event_date):
    ws = wb.active if first else wb.create_sheet()
    ws.title = key

    n = len(cols)
    last = get_column_letter(n)

    # ── Title band ────────────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{last}1')
    ws['A1'] = f'UTKARSH 2026  ·  {title}'
    ws['A1'].font = TITLE_F
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f'A2:{last}2')
    ws['A2'] = subtitle
    ws['A2'].font = SUB_F
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16

    # Judge/counter signature line — every paper sheet needs to say who held it.
    ws.merge_cells(f'A3:{last}3')
    ws['A3'] = 'Judge / Volunteer name: ______________________________        '
    ws['A3'].font = SUB_F
    ws.row_dimensions[3].height = 18

    HEAD_ROW = 5

    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=HEAD_ROW, column=i, value=c.title)
        cell.font = HEAD_F
        cell.fill = PatternFill('solid', fgColor=BAND)
        cell.alignment = WRAP_MID
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = c.width
        if c.note:
            cell.comment = None  # kept simple: notes live in the subtitle instead
    ws.row_dimensions[HEAD_ROW].height = 34

    # ── Body ──────────────────────────────────────────────────────────────────
    r = HEAD_ROW + 1
    for rec in rows:
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i, value=rec.get(c.title, ''))
            cell.font = CELL_F
            cell.border = BOX
            cell.alignment = WRAP_MID if c.align == 'center' else LEFT_MID
        ws.row_dimensions[r].height = 24
        r += 1

    spot_from = r
    for _ in range(spot_rows):
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=i)
            cell.font = CELL_F
            cell.border = BOX
            cell.fill = PatternFill('solid', fgColor=SPOT_FILL)
            cell.alignment = WRAP_MID if c.align == 'center' else LEFT_MID
        ws.row_dimensions[r].height = 24
        r += 1
    last_row = r - 1

    # Serial numbers run through the blanks too, so a half-filled sheet still
    # tells you how many students are on it.
    if cols and cols[0].title == 'S.No':
        for i, rr in enumerate(range(HEAD_ROW + 1, last_row + 1), start=1):
            ws.cell(row=rr, column=1, value=i).alignment = WRAP_MID

    # ── Print setup: this is a paper artefact first ───────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0          # as many pages tall as it needs
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f'{HEAD_ROW}:{HEAD_ROW}'
    ws.page_margins.left = ws.page_margins.right = 0.3
    ws.page_margins.top = ws.page_margins.bottom = 0.4
    ws.freeze_panes = ws.cell(row=HEAD_ROW + 1, column=1)
    ws.auto_filter.ref = f'A{HEAD_ROW}:{last}{last_row}'
    ws.oddFooter.right.text = 'Page &P of &N'
    ws.oddFooter.left.text = f'Utkarsh 2026 · {title} · {event_date}'
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.size = 8
    return ws, HEAD_ROW, spot_from, last_row


def total_formula(ws, head_row, first_row, last_row, score_cols, total_col):
    """Live SUM in the Total column, so a corrected score re-totals itself."""
    a = get_column_letter(score_cols[0])
    b = get_column_letter(score_cols[-1])
    for r in range(first_row, last_row + 1):
        c = ws.cell(row=r, column=total_col, value=f'=SUM({a}{r}:{b}{r})')
        c.font = Font(name='Calibri', size=10, bold=True, color=INK)
        c.alignment = WRAP_MID
        c.border = BOX


def yes_no(ws, head_row, first_row, last_row, col):
    dv = DataValidation(type='list', formula1='"YES,NO"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'{get_column_letter(col)}{first_row}:{get_column_letter(col)}{last_row}')


# ── Column sets ───────────────────────────────────────────────────────────────
# The three class bands the competitions are actually run and judged in. Each
# gets its own tab, so a spot registration is added to the sheet its group is
# already on rather than at the bottom of one long mixed list.
GROUPS = [('A', 'Class 1 to 4', 1, 4),
          ('B', 'Class 5 to 7', 5, 7),
          ('C', 'Class 8 to 10', 8, 10)]


def group_of(row):
    """Prefer the group the portal recorded; fall back to the class number."""
    g = str(row.get('Group') or '').strip().upper()
    if g in ('A', 'B', 'C'):
        return g
    try:
        c = int(str(row.get('Class') or '').strip())
    except ValueError:
        return None
    for code, _, lo, hi in GROUPS:
        if lo <= c <= hi:
            return code
    return None


IDENT = lambda: [
    Col('S.No', 5, align='center'),
    Col('Reg Code', 11, align='center'),
    Col('Name', 21),
    Col('Class', 6, align='center'),
    Col('Group', 6, align='center'),
    Col('School', 21),
    Col('Age', 5, align='center'),
]

SHLOKA_CRITERIA = [
    ('Pronunciation\n(Uchcharana)\n/10', 11),
    ('Memory &\nAccuracy\n/10', 10),
    ('Tone & Melody\n(Svara)\n/10', 11),
    ('Confidence &\nPresentation\n/10', 11),
    ('Overall\nImpression\n/10', 10),
]

ESSAY_CRITERIA = [
    ('Content &\nUnderstanding\n/10', 12),
    ('Language &\nGrammar\n/10', 11),
    ('Structure &\nFlow\n/10', 10),
    ('Originality &\nExpression\n/10', 11),
    ('Handwriting &\nNeatness\n/10', 11),
]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--csv', nargs='?', const='auto',
                    help='registrations export from the admin portal; '
                         'pass --csv with no path to pick the newest one in ~/Downloads')
    ap.add_argument('--date', default='23 August 2026', help='date printed on the sheets')
    ap.add_argument('--spot', type=int, default=25, help='blank rows for spot registrations')
    ap.add_argument('--out', default='sheets', help='output directory')
    args = ap.parse_args()

    today = dt.date(2026, 8, 23)
    out = pathlib.Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    # Finding the export by hand is the fiddly step on a busy morning, so
    # --csv with no path just takes the newest likely-looking one.
    if args.csv == 'auto':
        downloads = pathlib.Path.home() / 'Downloads'
        found = sorted(downloads.glob('*.csv'), key=lambda f: f.stat().st_mtime, reverse=True)
        found = [f for f in found if re.search(r'reg|utkarsh|export', f.name, re.I)] or found
        if not found:
            sys.exit(f'No CSV found in {downloads}. Export one from the admin portal first.')
        args.csv = str(found[0])
        print(f'using {args.csv}')

    records = []
    if args.csv:
        with open(args.csv, newline='', encoding='utf-8-sig') as fh:
            records = list(csv.DictReader(fh))
        if not records:
            print('WARNING: that CSV has no data rows — the sheets will come out blank.')
        else:
            missing = [c for c in ('Name', 'Class', 'School') if c not in records[0]]
            if missing:
                print(f'WARNING: the CSV is missing columns {missing} — re-export with all '
                      f'columns ticked, or those cells will be empty.')

    def get(rec, *names):
        for n in names:
            if n in rec and str(rec[n]).strip() not in ('', '—'):
                return str(rec[n]).strip()
        return ''

    def entered(rec, track_label):
        return get(rec, track_label) != ''

    def ident_rows(track_label=None, detail_col=None):
        """
        One row per student entered in `track_label`.

        `detail_col` carries the student's own choice — which shloka, which
        essay topic — through to the judge, who otherwise has to ask each
        child at the microphone. The export writes a bare 'Yes' when there is
        no choice stored, which is not a shloka, so it is dropped.
        """
        rows = []
        for rec in records:
            if track_label and not entered(rec, track_label):
                continue
            dob = parse_date(get(rec, 'Date of birth'))
            row = {
                'Reg Code': get(rec, 'Registration code'),
                'Name': get(rec, 'Name'),
                'Class': get(rec, 'Class'),
                'Group': get(rec, 'Group'),
                'School': get(rec, 'School'),
                'Age': age_on(dob, today),
                'Date of birth': get(rec, 'Date of birth'),
            }
            if detail_col and track_label:
                choice = get(rec, track_label)
                row[detail_col] = '' if choice.lower() == 'yes' else choice
            rows.append(row)
        rows.sort(key=lambda r: (str(r['Group']), str(r['Name']).lower()))
        return rows

    made = []

    def judging_workbook(filename, key_prefix, title, subtitle, cols, rows,
                         score_cols=None, total_col=None, present_col=None):
        """
        One workbook, one tab per group, same columns on each.

        Anyone whose group cannot be worked out still gets a tab of their own
        rather than being dropped — a missing student is far worse than an
        untidy workbook.
        """
        by_group = {code: [] for code, _, _, _ in GROUPS}
        stray = []
        for r in rows:
            g = group_of(r)
            (by_group[g] if g else stray).append(r)

        wb = Workbook()
        first = True
        counts = []
        tabs = [(code, label, by_group[code]) for code, label, _, _ in GROUPS]
        if stray:
            tabs.append(('?', 'group not recorded', stray))

        for code, label, grp_rows in tabs:
            ws, hr, spot_from, last_row = build_sheet(
                wb, key=f'Group {code}',
                title=f'{title}  ·  Group {code}',
                subtitle=f'{label}  ·  {subtitle}',
                cols=cols, rows=grp_rows, spot_rows=args.spot,
                first=first, event_date=args.date)
            if score_cols:
                total_formula(ws, hr, hr + 1, last_row, score_cols, total_col)
            if present_col:
                yes_no(ws, hr, hr + 1, last_row, present_col)
            counts.append((code, len(grp_rows)))
            first = False

        path = out / filename
        wb.save(path)
        made.append((path, counts))
        return path

    # ── 1. Gita Shloka ────────────────────────────────────────────────────────
    cols = IDENT() + [Col('Shloka(s)\nrecited', 15)]
    score_from = len(cols) + 1
    cols += [Col(t, w, fill_in=True, align='center') for t, w in SHLOKA_CRITERIA]
    score_to = len(cols)
    cols += [Col('TOTAL\n/50', 9, align='center'), Col('Rank', 7, align='center'),
             Col('Judge remarks', 22)]
    judging_workbook(
        'gita-shloka-judging.xlsx', 'Gita Shloka',
        'Gita Shloka Recitation — Judging Sheet',
        f'{args.date}  ·  each criterion out of 10, total out of 50'
        f'  ·  shaded rows are for spot registrations',
        cols, ident_rows('Gita Shloka Recitation', 'Shloka(s)\nrecited'),
        score_cols=list(range(score_from, score_to + 1)), total_col=score_to + 1)

    # ── 2. Vedic Quiz ─────────────────────────────────────────────────────────
    cols = IDENT() + [
        Col('Present', 9, align='center'),
        Col('Round 1\n/15', 9, align='center'),
        Col('Round 2\n/15', 9, align='center'),
        Col('TOTAL\n/30', 9, align='center'),
        Col('Rank', 7, align='center'),
        Col('Remarks', 24),
    ]
    r1 = len(IDENT()) + 2
    judging_workbook(
        'vedic-quiz-marks.xlsx', 'Vedic Quiz', 'Vedic Quiz — Marks Sheet',
        f'{args.date}  ·  two rounds of 15 questions  ·  scores come from the quiz'
        f' platform; this sheet is the paper record  ·  shaded rows are for spot registrations',
        cols, ident_rows('Vedic Quiz'),
        score_cols=[r1, r1 + 1], total_col=r1 + 2, present_col=len(IDENT()) + 1)

    # ── 3. Devotional Essay ───────────────────────────────────────────────────
    cols = IDENT() + [Col('Topic chosen', 20)]
    score_from = len(cols) + 1
    cols += [Col(t, w, fill_in=True, align='center') for t, w in ESSAY_CRITERIA]
    score_to = len(cols)
    cols += [Col('TOTAL\n/50', 9, align='center'), Col('Rank', 7, align='center'),
             Col('Judge remarks', 22)]
    judging_workbook(
        'devotional-essay-judging.xlsx', 'Devotional Essay',
        'Devotional Essay — Judging Sheet',
        f'{args.date}  ·  each criterion out of 10, total out of 50'
        f'  ·  shaded rows are for spot registrations',
        cols, ident_rows('Devotional Essay', 'Topic chosen'),
        score_cols=list(range(score_from, score_to + 1)), total_col=score_to + 1)

    # ── 4. Registration counter ───────────────────────────────────────────────
    TRACKS = ['Vedic Quiz', 'Gita Shloka Recitation', 'Devotional Essay',
              'Vedic Art', 'Vedic Fancy Dress', 'Devotional Bhajan']
    cols = [
        Col('S.No', 6, align='center'), Col('Reg Code', 13, align='center'),
        Col('Name', 24), Col('Class', 7, align='center'), Col('Group', 7, align='center'),
        Col('School', 26), Col('Age', 6, align='center'), Col('Gender', 8, align='center'),
        Col('Guardian', 20), Col('Guardian phone', 15, align='center'),
        Col('WhatsApp', 15, align='center'), Col('Email', 24),
    ]
    track_from = len(cols) + 1
    cols += [Col(t.replace(' ', '\n'), 11, align='center') for t in TRACKS]
    track_to = len(cols)
    cols += [
        Col('No. of\nevents', 9, align='center'),
        Col('Fee due\n(₹99 each)', 11, align='center'),
        Col('Payment\nstatus', 11, align='center'),
        Col('Amount\nreceived', 11, align='center'),
        Col('Receipt /\nUPI ref', 15, align='center'),
        Col('Collected\nby', 13, align='center'),
        Col('Remarks', 24),
    ]

    rows = []
    for rec in records:
        dob = parse_date(get(rec, 'Date of birth'))
        row = {
            'Reg Code': get(rec, 'Registration code'), 'Name': get(rec, 'Name'),
            'Class': get(rec, 'Class'), 'Group': get(rec, 'Group'),
            'School': get(rec, 'School'), 'Age': age_on(dob, today),
            'Gender': get(rec, 'Gender'), 'Guardian': get(rec, 'Guardian'),
            'Guardian phone': get(rec, 'Guardian phone'),
            'WhatsApp': get(rec, 'WhatsApp'), 'Email': get(rec, 'Email'),
            'Payment\nstatus': 'PAID' if get(rec, 'Paid') == 'YES' else 'PENDING',
            'Receipt /\nUPI ref': get(rec, 'Payment ref'),
        }
        for t in TRACKS:
            row[t.replace(' ', '\n')] = 'YES' if entered(rec, t) else ''
        rows.append(row)
    rows.sort(key=lambda r: str(r['Name']).lower())

    wb = Workbook()
    ws, hr, spot_from, last_row = build_sheet(
        wb, key='Registration counter', title='Registration Counter — Master Sheet',
        subtitle=f'{args.date}  ·  ₹99 per competition  ·  tick each event entered, then fill fee and payment'
                 f'  ·  shaded rows are for spot registrations',
        cols=cols, rows=rows, spot_rows=max(args.spot * 2, 50), first=True, event_date=args.date)

    # Events counted and fee derived from the ticks, so the counter never has to
    # multiply ₹99 by hand under pressure.
    a, b = get_column_letter(track_from), get_column_letter(track_to)
    n_col, fee_col = track_to + 1, track_to + 2
    for r in range(hr + 1, last_row + 1):
        c = ws.cell(row=r, column=n_col, value=f'=COUNTIF({a}{r}:{b}{r},"YES")')
        c.font, c.alignment, c.border = CELL_F, WRAP_MID, BOX
        f = ws.cell(row=r, column=fee_col, value=f'=IF({get_column_letter(n_col)}{r}=0,"",{get_column_letter(n_col)}{r}*99)')
        f.font = Font(name='Calibri', size=10, bold=True, color=INK)
        f.alignment, f.border = WRAP_MID, BOX

    dv = DataValidation(type='list', formula1='"PAID,PENDING"', allow_blank=True)
    ws.add_data_validation(dv)
    ps = get_column_letter(track_to + 3)
    dv.add(f'{ps}{hr+1}:{ps}{last_row}')
    for t in range(track_from, track_to + 1):
        yes_no(ws, hr, hr + 1, last_row, t)

    ws.page_setup.fitToWidth = 2
    ws.print_title_cols = 'A:C'

    p = out / 'registration-counter.xlsx'; wb.save(p)
    made.append((p, [('all', len(rows))]))

    print(f'students from CSV: {len(records)}' if args.csv else 'no CSV given — blank sheets')
    for path, counts in made:
        detail = '  '.join(f'{c}:{n}' for c, n in counts)
        print(f'  {path.name:<32} {sum(n for _, n in counts):>3} students   {detail}')
    if args.csv:
        # A student who reaches no sheet is the one failure that matters here.
        placed = {c for _, counts in made for c, _ in counts}
        if '?' in placed:
            print('\n  NOTE: some students had no group recorded and are on a '
                  '"Group ?" tab — check their class.')


if __name__ == '__main__':
    main()
